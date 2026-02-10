from __future__ import annotations

from pathlib import Path
from typing import Optional

import tkinter as tk
from tkinter import filedialog, messagebox, simpledialog, ttk

from ...database import Database
from ...models import ProductionLogClient, ProductionLogSheetConfig
from ...theme import ThemePalette

try:  # Optional dependency for reading Excel files.
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore
except Exception:  # pragma: no cover - handled at runtime
    load_workbook = None
    get_column_letter = None


FIELD_LABELS = {
    "run_date": "Run Date",
    "dry_file_name": "Dry - File Name",
    "live_file_name": "Live - File Name",
    "dry_dist_id": "Dry - Dist ID",
    "live_dist_id": "Live - Dist ID",
    "dry_run_counts": "Dry Run Counts",
    "live_run_counts": "Live Run Counts",
    "system_a_count_hyphen": "System-A Count",
    "system_a_pkg_code_hyphen": "System-A PKG Code",
    "lettershop_recd": "Lettershop REC'd",
    "comments": "Comments",
    "mass_cancel": "Mass Cancel",
    "coreqc_email": "CoreQC-email",
    "check_sl": "Check SL",
    "send_approval_engage": "Send Approval to Engage",
    "dry_report_name": "Dry - Report Name",
    "live_report_name": "Live - Report Name",
    "dry_vs_live_run_counts": "Dry Run Counts vs Live Run Counts",
    "dist_file_name": "Dist File Name",
    "dist_run_id": "Dist Run ID",
    "cpn330": "CPN330",
    "csv_counts": "CSV Counts",
    "smc_initialization": "SMC Initialization",
    "email_mktg_srv": "Email Mktg Srv",
    "system_a_counts": "System A counts",
    "system_a_count_plain": "System A Count",
    "system_a_pkg_code_plain": "System A PKG Code",
    "send_approval_email": "Send Approval Email",
    "job_ran_successfully": "Job Ran Successfully",
    "counts": "Counts",
    "errors": "Errors",
    "in_archive": "In Archive",
    "dist_id": "Dist ID",
    "dist_file": "Dist File",
    "records_exported": "Records Exported",
    "output_recd": "Output REC'd",
    "in_cdsfiles": "In CDSFiles",
    "error_comments": "Error Comments",
    "select_set": "Select Set",
    "jobstream": "Jobstream",
}

SHEET_TEMPLATES: list[tuple[str, str, list[str]]] = [
    (
        "single_bills_mailed",
        "Single Bills - Mailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "system_a_count_hyphen",
            "system_a_pkg_code_hyphen",
            "lettershop_recd",
            "comments",
            "mass_cancel",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "single_bills_emailed",
        "Single Bills - Emailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "coreqc_email",
            "check_sl",
            "send_approval_engage",
            "comments",
            "mass_cancel",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "gift_bills_mailed",
        "Gift Bills - Mailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "system_a_count_hyphen",
            "system_a_pkg_code_hyphen",
            "lettershop_recd",
            "comments",
            "mass_cancel",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "gift_bills_emailed",
        "Gift Bills - Emailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "coreqc_email",
            "check_sl",
            "send_approval_engage",
            "comments",
            "mass_cancel",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "single_renewals_mailed",
        "Single Renewals - Mailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "system_a_count_hyphen",
            "system_a_pkg_code_hyphen",
            "lettershop_recd",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "single_renewals_emailed",
        "Single Renewals - Emailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "coreqc_email",
            "check_sl",
            "send_approval_engage",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "single_unren_renewals_mailed",
        "Single UnRen Renewals - Mailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "system_a_count_hyphen",
            "system_a_pkg_code_hyphen",
            "lettershop_recd",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "single_unren_renewals_emailed",
        "Single UnRen Renewals - Emailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "coreqc_email",
            "check_sl",
            "send_approval_engage",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "auto_renewal_links_mailed",
        "Auto Renewal Links - Mailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "system_a_count_hyphen",
            "system_a_pkg_code_hyphen",
            "lettershop_recd",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "auto_renewal_links_emailed",
        "Auto Renewal Links - Emailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "coreqc_email",
            "check_sl",
            "send_approval_engage",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "auto_renewal_apply",
        "Auto Renewal Apply",
        [
            "dry_run_counts",
            "live_run_counts",
            "dry_dist_id",
            "live_dist_id",
            "dry_report_name",
            "live_report_name",
            "comments",
            "dry_vs_live_run_counts",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "gift_renewal_mailed",
        "Gift Renewal - Mailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "system_a_count_hyphen",
            "system_a_pkg_code_hyphen",
            "lettershop_recd",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "gift_renewal_emailed",
        "Gift Renewal - Emailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "coreqc_email",
            "check_sl",
            "send_approval_engage",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "gift_unren_renewal_mailed",
        "Gift UnRen Renewal - Mailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "system_a_count_hyphen",
            "system_a_pkg_code_hyphen",
            "lettershop_recd",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "gift_unren_renewal_emailed",
        "Gift UnRen Renewal - Emailed",
        [
            "dry_file_name",
            "live_file_name",
            "dry_dist_id",
            "live_dist_id",
            "dry_run_counts",
            "live_run_counts",
            "run_date",
            "coreqc_email",
            "check_sl",
            "send_approval_engage",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "ancillary_acks",
        "Ancillary Acks",
        [
            "dist_file_name",
            "dist_run_id",
            "run_date",
            "cpn330",
            "csv_counts",
            "smc_initialization",
            "email_mktg_srv",
            "system_a_counts",
            "lettershop_recd",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "cold_donor_mailed",
        "Cold Donor - Mailed",
        [
            "dist_file_name",
            "dist_run_id",
            "run_date",
            "system_a_count_plain",
            "system_a_pkg_code_plain",
            "lettershop_recd",
            "comments",
        ],
    ),
    (
        "cold_donor_emailed",
        "Cold Donor - Emailed",
        [
            "dist_file_name",
            "dist_run_id",
            "run_date",
            "coreqc_email",
            "check_sl",
            "send_approval_email",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "recurring_donations",
        "Recurring Donations",
        [
            "run_date",
            "job_ran_successfully",
            "counts",
            "errors",
            "in_archive",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "premium_ship",
        "Premium Ship",
        [
            "run_date",
            "in_archive",
            "dist_id",
            "job_ran_successfully",
            "counts",
            "errors",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "single_copy_ship",
        "Single Copy Ship",
        [
            "run_date",
            "in_archive",
            "dist_id",
            "job_ran_successfully",
            "counts",
            "errors",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "cycle_end",
        "Cycle End",
        [
            "run_date",
            "dist_file",
            "dist_run_id",
            "records_exported",
            "output_recd",
            "system_a_count_plain",
            "system_a_pkg_code_plain",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
    (
        "eod",
        "EOD",
        [
            "run_date",
            "job_ran_successfully",
            "errors",
            "in_archive",
            "in_cdsfiles",
            "error_comments",
            "comments",
            "select_set",
            "jobstream",
        ],
    ),
]

TEMPLATE_LABELS = {key: label for key, label, _fields in SHEET_TEMPLATES}
TEMPLATE_FIELDS = {key: fields for key, _label, fields in SHEET_TEMPLATES}
TEMPLATE_KEY_BY_LABEL = {label: key for key, label, _fields in SHEET_TEMPLATES}
TEMPLATE_PLACEHOLDER = "Select sheet type..."


class ProductionLogView(ttk.Frame):
    _PIN_CODE = "12345"
    _HEADER_ROW = 5
    _DATA_START_ROW = 6
    _PREVIEW_ROWS = 10

    def __init__(self, master: tk.Misc, db: Database, theme: ThemePalette) -> None:
        super().__init__(master, padding=(16, 16))
        self.db = db
        self.theme = theme
        self._locked = True
        self._lock_overlay: Optional[tk.Frame] = None
        self._pin_entry: Optional[ttk.Entry] = None
        self._pin_var = tk.StringVar(value="")
        self._lock_error_var = tk.StringVar(value="")
        self._accent_strip: Optional[tk.Frame] = None

        self.clients: list[ProductionLogClient] = []
        self.current_client_id: Optional[int] = None
        self.sheet_configs: dict[str, ProductionLogSheetConfig] = {}
        self._column_choice_map: dict[str, str] = {}
        self._last_column_choices: list[str] = []
        self._active_template_key: Optional[str] = None
        self._active_field_keys: list[str] = []

        self.client_var = tk.StringVar(value="")
        self.workbook_var = tk.StringVar(value="")
        self.sheet_var = tk.StringVar(value="")
        self.template_var = tk.StringVar(value=TEMPLATE_PLACEHOLDER)
        self.status_var = tk.StringVar(value="Read-only preview mode.")

        self._field_vars: dict[str, tk.StringVar] = {
            key: tk.StringVar(value="") for key in FIELD_LABELS
        }
        self._mapping_inputs: dict[str, ttk.Combobox] = {}
        self._mapping_field_frame: Optional[ttk.Frame] = None

        self._configure_styles()
        self.configure(style="ProdLog.Root.TFrame")
        self._build_ui()
        self._load_clients()
        self.after(0, self._show_lock_overlay)

    # ------------------------------------------------------------------ UI
    def _build_ui(self) -> None:
        hero = ttk.Frame(self, style="ProdLog.Hero.TFrame", padding=(16, 12))
        hero.pack(fill=tk.X)
        accent = tk.Frame(hero, width=4, bg=self.theme.accent)
        accent.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 12))
        self._accent_strip = accent
        hero_text = ttk.Frame(hero, style="ProdLog.Hero.TFrame")
        hero_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        ttk.Label(hero_text, text="Production Log", style="ProdLog.Title.TLabel").pack(anchor="w")
        ttk.Label(
            hero_text,
            text=(
                "Link a client workbook to preview and map production log columns. "
                "Updates will run only when the workbook is not in use."
            ),
            style="ProdLog.Body.TLabel",
            wraplength=860,
            justify="left",
        ).pack(anchor="w", pady=(6, 0))
        ttk.Label(hero, textvariable=self.status_var, style="ProdLog.Badge.TLabel").pack(
            side=tk.RIGHT, padx=(12, 0)
        )

        body = ttk.Frame(self, style="ProdLog.Root.TFrame")
        body.pack(fill=tk.BOTH, expand=True, pady=(14, 0))

        paned = ttk.Panedwindow(body, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)

        left = ttk.Frame(paned, style="ProdLog.Root.TFrame")
        left.columnconfigure(0, weight=1)

        right = ttk.Frame(paned, style="ProdLog.Root.TFrame")
        right.columnconfigure(0, weight=1)

        paned.add(left, weight=1)
        paned.add(right, weight=2)

        client_card = ttk.Frame(left, style="ProdLog.Card.TFrame", padding=(16, 14))
        client_card.pack(fill=tk.X, pady=(0, 12))
        ttk.Label(client_card, text="Client Setup", style="ProdLog.Section.TLabel").pack(anchor="w")

        client_row = ttk.Frame(client_card, style="ProdLog.Card.TFrame")
        client_row.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(client_row, text="Client", style="ProdLog.Card.TLabel").pack(side=tk.LEFT)
        self.client_combo = ttk.Combobox(client_row, textvariable=self.client_var, state="readonly", width=28)
        self.client_combo.pack(side=tk.LEFT, padx=(8, 6))
        self.client_combo.bind("<<ComboboxSelected>>", self._on_client_selected)
        ttk.Button(client_row, text="New...", command=self._new_client).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(client_row, text="Rename...", command=self._rename_client).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(client_row, text="Delete", command=self._delete_client).pack(side=tk.LEFT)

        workbook_row = ttk.Frame(client_card, style="ProdLog.Card.TFrame")
        workbook_row.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(workbook_row, text="Workbook", style="ProdLog.Card.TLabel").pack(side=tk.LEFT)
        self.workbook_entry = ttk.Entry(
            workbook_row,
            textvariable=self.workbook_var,
            state="readonly",
            width=54,
        )
        self.workbook_entry.pack(side=tk.LEFT, padx=(8, 6), fill=tk.X, expand=True)
        ttk.Button(workbook_row, text="Select Workbook...", command=self._choose_workbook).pack(side=tk.LEFT)

        sheet_row = ttk.Frame(client_card, style="ProdLog.Card.TFrame")
        sheet_row.pack(fill=tk.X, pady=(10, 0))
        ttk.Label(sheet_row, text="Sheet", style="ProdLog.Card.TLabel").pack(side=tk.LEFT)
        self.sheet_combo = ttk.Combobox(sheet_row, textvariable=self.sheet_var, state="readonly", width=24)
        self.sheet_combo.pack(side=tk.LEFT, padx=(8, 6))
        self.sheet_combo.bind("<<ComboboxSelected>>", self._on_sheet_selected)
        ttk.Button(sheet_row, text="Refresh Sheets", command=self._load_workbook_sheets).pack(side=tk.LEFT)

        mapping_card = ttk.Frame(left, style="ProdLog.Card.TFrame", padding=(16, 14))
        mapping_card.pack(fill=tk.BOTH, expand=True)
        mapping_card.columnconfigure(1, weight=1)
        ttk.Label(
            mapping_card,
            text=f"Column Mapping (header row {self._HEADER_ROW}, data starts row {self._DATA_START_ROW})",
            style="ProdLog.Section.TLabel",
        ).grid(row=0, column=0, columnspan=2, sticky="w")
        ttk.Label(
            mapping_card,
            text="Map the columns you want to fill later. Nothing is written yet.",
            style="ProdLog.BodyMuted.TLabel",
        ).grid(row=1, column=0, columnspan=2, sticky="w", pady=(4, 8))

        ttk.Label(mapping_card, text="Sheet Type", style="ProdLog.Card.TLabel").grid(
            row=2, column=0, sticky="w", pady=2, padx=(0, 8)
        )
        self.template_combo = ttk.Combobox(
            mapping_card,
            textvariable=self.template_var,
            state="readonly",
            width=28,
        )
        self.template_combo["values"] = [TEMPLATE_PLACEHOLDER] + [
            label for _key, label, _fields in SHEET_TEMPLATES
        ]
        self.template_combo.grid(row=2, column=1, sticky="w", pady=2)
        self.template_combo.bind("<<ComboboxSelected>>", self._on_template_selected)

        self._mapping_field_frame = ttk.Frame(mapping_card, style="ProdLog.Card.TFrame")
        self._mapping_field_frame.grid(row=3, column=0, columnspan=2, sticky="nsew", pady=(6, 0))
        self._mapping_field_frame.columnconfigure(1, weight=1)
        self._render_mapping_fields()

        button_row = ttk.Frame(mapping_card, style="ProdLog.Card.TFrame")
        button_row.grid(row=4, column=0, columnspan=2, sticky="e", pady=(10, 0))
        ttk.Button(button_row, text="Clear Mapping", command=self._clear_mapping).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(button_row, text="Save Mapping", command=self._save_mapping).pack(side=tk.LEFT)

        right_pane = ttk.Panedwindow(right, orient=tk.VERTICAL)
        right_pane.pack(fill=tk.BOTH, expand=True)

        preview_card = ttk.Frame(right_pane, style="ProdLog.Card.TFrame", padding=(16, 14))
        preview_card.columnconfigure(0, weight=1)
        preview_card.rowconfigure(1, weight=1)
        ttk.Label(preview_card, text="Preview (read-only)", style="ProdLog.Section.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        self.preview_tree = ttk.Treeview(preview_card, show="headings", height=12, style="ProdLog.Treeview")
        self.preview_tree.grid(row=1, column=0, sticky="nsew", pady=(8, 0))
        scroll = ttk.Scrollbar(preview_card, orient=tk.VERTICAL, command=self.preview_tree.yview)
        scroll.grid(row=1, column=1, sticky="ns", pady=(8, 0))
        scroll_x = ttk.Scrollbar(preview_card, orient=tk.HORIZONTAL, command=self.preview_tree.xview)
        scroll_x.grid(row=2, column=0, sticky="ew", pady=(6, 0))
        self.preview_tree.configure(yscrollcommand=scroll.set, xscrollcommand=scroll_x.set)
        ttk.Label(
            preview_card,
            text="Showing the first 10 rows starting at row 6.",
            style="ProdLog.BodyMuted.TLabel",
        ).grid(row=3, column=0, sticky="w", pady=(8, 0))

        preview_spacer = ttk.Frame(right_pane, style="ProdLog.Root.TFrame")
        right_pane.add(preview_card, weight=1)
        right_pane.add(preview_spacer, weight=0)

    # ------------------------------------------------------------------ Client management
    def _load_clients(self) -> None:
        self.clients = self.db.get_production_log_clients()
        names = [client.name for client in self.clients]
        self.client_combo["values"] = names
        if self.current_client_id and any(c.id == self.current_client_id for c in self.clients):
            current = next(c for c in self.clients if c.id == self.current_client_id)
            self.client_combo.set(current.name)
        elif names:
            self.client_combo.current(0)
            self.current_client_id = self.clients[0].id
        else:
            self.client_combo.set("")
            self.current_client_id = None
        self._sync_client_state()

    def _sync_client_state(self) -> None:
        if self.current_client_id is None:
            self.workbook_var.set("")
            self.sheet_combo["values"] = []
            self.sheet_var.set("")
            self.sheet_configs = {}
            self._update_mapping_inputs([])
            self._clear_preview()
            self.status_var.set("Select a client to begin.")
            return
        current = next((c for c in self.clients if c.id == self.current_client_id), None)
        if current:
            self.workbook_var.set(current.workbook_path or "")
        self.sheet_configs = {
            cfg.sheet_name: cfg
            for cfg in self.db.get_production_log_sheet_configs(self.current_client_id)
        }
        self._refresh_status()
        self._load_workbook_sheets()

    def _on_client_selected(self, _event: object) -> None:
        name = self.client_var.get()
        match = next((c for c in self.clients if c.name == name), None)
        if match:
            self.current_client_id = match.id
        self._sync_client_state()

    def _new_client(self) -> None:
        name = simpledialog.askstring("New Client", "Client name:", parent=self)
        if not name:
            return
        try:
            new_id = self.db.create_production_log_client(name)
        except ValueError as exc:
            messagebox.showerror("Client", str(exc), parent=self)
            return
        self.current_client_id = new_id
        self._load_clients()

    def _rename_client(self) -> None:
        if self.current_client_id is None:
            return
        current = next((c for c in self.clients if c.id == self.current_client_id), None)
        if current is None:
            return
        name = simpledialog.askstring("Rename Client", "New name:", initialvalue=current.name, parent=self)
        if not name:
            return
        try:
            self.db.update_production_log_client(self.current_client_id, name)
        except ValueError as exc:
            messagebox.showerror("Client", str(exc), parent=self)
            return
        self._load_clients()

    def _delete_client(self) -> None:
        if self.current_client_id is None:
            return
        current = next((c for c in self.clients if c.id == self.current_client_id), None)
        if current is None:
            return
        confirm = messagebox.askyesno(
            "Delete Client",
            f"Delete client '{current.name}' and its mappings? This cannot be undone.",
            parent=self,
        )
        if not confirm:
            return
        self.db.delete_production_log_client(self.current_client_id)
        self.current_client_id = None
        self._load_clients()

    # ------------------------------------------------------------------ Workbook handling
    def _choose_workbook(self) -> None:
        if self.current_client_id is None:
            messagebox.showinfo("Workbook", "Create or select a client first.", parent=self)
            return
        path = filedialog.askopenfilename(
            parent=self,
            title="Select Production Log Workbook",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return
        self.db.update_production_log_client_workbook(self.current_client_id, path)
        self._load_clients()

    def _refresh_status(self) -> None:
        path_text = (self.workbook_var.get() or "").strip()
        if not path_text:
            self.status_var.set("Read-only preview mode.")
            return
        path = Path(path_text)
        if not path.exists():
            self.status_var.set("Workbook not found.")
            return
        lock_path = path.parent / f"~${path.name}"
        if lock_path.exists():
            self.status_var.set("In use by another user. Updates paused.")
        else:
            self.status_var.set("Workbook available.")

    def _load_workbook_sheets(self) -> None:
        self._refresh_status()
        path_text = (self.workbook_var.get() or "").strip()
        if not path_text or not Path(path_text).exists():
            self.sheet_combo["values"] = []
            self.sheet_var.set("")
            self._update_mapping_inputs([])
            self._clear_preview()
            return
        if load_workbook is None:
            self.status_var.set("Install openpyxl to load workbook sheets.")
            return
        try:
            workbook = load_workbook(path_text, read_only=True, data_only=True)
        except Exception as exc:
            self.status_var.set(f"Unable to read workbook: {exc}")
            return
        try:
            sheets = list(workbook.sheetnames)
        finally:
            workbook.close()
        self.sheet_combo["values"] = sheets
        current = self.sheet_var.get()
        if current not in sheets:
            if sheets:
                self.sheet_var.set(sheets[0])
            else:
                self.sheet_var.set("")
        self._on_sheet_selected()

    def _on_sheet_selected(self, _event: object | None = None) -> None:
        sheet_name = self.sheet_var.get().strip()
        if not sheet_name:
            self._update_mapping_inputs([])
            self._clear_preview()
            return
        self._load_column_choices(sheet_name)
        self._load_sheet_mapping(sheet_name)
        self._load_sheet_preview(sheet_name)

    def _load_column_choices(self, sheet_name: str) -> None:
        path_text = (self.workbook_var.get() or "").strip()
        self._column_choice_map = {}
        self._last_column_choices = []
        if not path_text or load_workbook is None:
            self._update_mapping_inputs([])
            return
        try:
            workbook = load_workbook(path_text, read_only=True, data_only=True)
        except Exception:
            self._update_mapping_inputs([])
            return
        try:
            if sheet_name not in workbook.sheetnames:
                self._update_mapping_inputs([])
                return
            sheet = workbook[sheet_name]
            header_rows = list(
                sheet.iter_rows(
                    min_row=self._HEADER_ROW,
                    max_row=self._HEADER_ROW,
                    values_only=True,
                )
            )
            header_values = list(header_rows[0]) if header_rows else []
            preview_rows = list(
                sheet.iter_rows(
                    min_row=self._DATA_START_ROW,
                    max_row=self._DATA_START_ROW + self._PREVIEW_ROWS - 1,
                    values_only=True,
                )
            )
            max_col = 0
            for idx, header in enumerate(header_values, start=1):
                if header not in (None, ""):
                    max_col = max(max_col, idx)
            for row in preview_rows:
                for idx, cell in enumerate(row, start=1):
                    if cell not in (None, ""):
                        max_col = max(max_col, idx)
            if max_col == 0:
                max_col = max(len(header_values), 1)
            choices = []
            for col_idx in range(1, max_col + 1):
                letter = get_column_letter(col_idx) if get_column_letter else str(col_idx)
                header_text = ""
                if col_idx - 1 < len(header_values):
                    header_text = str(header_values[col_idx - 1]).strip() if header_values[col_idx - 1] is not None else ""
                display = f"{letter} - {header_text}" if header_text else letter
                self._column_choice_map[display] = letter
                choices.append(display)
        finally:
            workbook.close()
        self._last_column_choices = choices
        self._update_mapping_inputs(choices)

    def _update_mapping_inputs(self, choices: list[str]) -> None:
        values = [""] + choices if choices else [""]
        for combo in self._mapping_inputs.values():
            combo.configure(values=values)

    def _render_mapping_fields(self) -> None:
        if self._mapping_field_frame is None:
            return
        for child in self._mapping_field_frame.winfo_children():
            child.destroy()
        self._mapping_inputs = {}
        if not self._active_field_keys:
            ttk.Label(
                self._mapping_field_frame,
                text="Select a sheet type to see the available fields.",
                style="ProdLog.BodyMuted.TLabel",
            ).grid(row=0, column=0, sticky="w")
            return
        for row_idx, key in enumerate(self._active_field_keys):
            label = FIELD_LABELS.get(key, key)
            ttk.Label(self._mapping_field_frame, text=label, style="ProdLog.Card.TLabel").grid(
                row=row_idx, column=0, sticky="w", pady=2, padx=(0, 8)
            )
            combo = ttk.Combobox(
                self._mapping_field_frame,
                textvariable=self._field_vars[key],
                state="readonly",
                width=26,
            )
            combo.grid(row=row_idx, column=1, sticky="w", pady=2)
            self._mapping_inputs[key] = combo
        self._update_mapping_inputs(self._last_column_choices)

    def _set_active_template(self, template_key: Optional[str]) -> None:
        self._active_template_key = template_key
        fields = list(TEMPLATE_FIELDS.get(template_key or "", []))
        priority = ["run_date", "select_set", "jobstream"]
        ordered = [field for field in priority if field in fields]
        ordered.extend([field for field in fields if field not in ordered])
        self._active_field_keys = ordered
        self._render_mapping_fields()

    def _select_template(self, template_key: Optional[str]) -> None:
        if template_key and template_key in TEMPLATE_LABELS:
            self.template_var.set(TEMPLATE_LABELS[template_key])
        else:
            self.template_var.set(TEMPLATE_PLACEHOLDER)
        self._set_active_template(template_key)

    def _on_template_selected(self, _event: object | None = None) -> None:
        label = self.template_var.get().strip()
        template_key = TEMPLATE_KEY_BY_LABEL.get(label)
        self._set_active_template(template_key)

    def _load_sheet_mapping(self, sheet_name: str) -> None:
        for var in self._field_vars.values():
            var.set("")
        self._select_template(None)
        if self.current_client_id is None:
            return
        config = self.sheet_configs.get(sheet_name)
        if config is None:
            config = self.db.get_production_log_sheet_config(self.current_client_id, sheet_name)
        if config is None:
            return
        self._select_template(config.template_key)
        for key, column in config.column_mappings.items():
            display = self._display_for_column(column)
            if key in self._field_vars:
                self._field_vars[key].set(display)

    def _display_for_column(self, column: str) -> str:
        column = (column or "").strip().upper()
        for display, letter in self._column_choice_map.items():
            if letter.upper() == column:
                return display
        return column

    def _clear_mapping(self) -> None:
        for var in self._field_vars.values():
            var.set("")

    def _save_mapping(self) -> None:
        if self.current_client_id is None:
            messagebox.showinfo("Mapping", "Select a client first.", parent=self)
            return
        sheet_name = self.sheet_var.get().strip()
        if not sheet_name:
            messagebox.showinfo("Mapping", "Select a sheet first.", parent=self)
            return
        if not self._active_template_key:
            messagebox.showinfo("Mapping", "Select a sheet type first.", parent=self)
            return
        mapping: dict[str, str] = {}
        duplicates: dict[str, list[str]] = {}
        used: dict[str, str] = {}
        for key in self._active_field_keys:
            var = self._field_vars[key]
            raw = var.get().strip()
            if not raw:
                continue
            column = self._column_choice_map.get(raw, raw).strip().upper()
            if column in used:
                duplicates.setdefault(column, [used[column]]).append(FIELD_LABELS.get(key, key))
            else:
                used[column] = FIELD_LABELS.get(key, key)
            mapping[key] = column
        if duplicates:
            details = []
            for column, fields in duplicates.items():
                fields_list = ", ".join(fields)
                details.append(f"{column}: {fields_list}")
            messagebox.showerror(
                "Mapping",
                "Each column can map to only one field.\n\nConflicts:\n" + "\n".join(details),
                parent=self,
            )
            return
        self.db.upsert_production_log_sheet_config(
            client_id=self.current_client_id,
            sheet_name=sheet_name,
            template_key=self._active_template_key,
            header_row=self._HEADER_ROW,
            data_start_row=self._DATA_START_ROW,
            column_mappings=mapping,
        )
        self.sheet_configs[sheet_name] = self.db.get_production_log_sheet_config(
            self.current_client_id, sheet_name
        ) or self.sheet_configs.get(sheet_name)
        messagebox.showinfo("Mapping", "Column mapping saved.", parent=self)

    # ------------------------------------------------------------------ Preview
    def _clear_preview(self) -> None:
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = []

    def _load_sheet_preview(self, sheet_name: str) -> None:
        self._clear_preview()
        path_text = (self.workbook_var.get() or "").strip()
        if not path_text or load_workbook is None:
            return
        try:
            workbook = load_workbook(path_text, read_only=True, data_only=True)
        except Exception:
            return
        try:
            if sheet_name not in workbook.sheetnames:
                return
            sheet = workbook[sheet_name]
            header_rows = list(
                sheet.iter_rows(
                    min_row=self._HEADER_ROW,
                    max_row=self._HEADER_ROW,
                    values_only=True,
                )
            )
            headers = list(header_rows[0]) if header_rows else []
            data_rows = list(
                sheet.iter_rows(
                    min_row=self._DATA_START_ROW,
                    max_row=self._DATA_START_ROW + self._PREVIEW_ROWS - 1,
                    values_only=True,
                )
            )
            max_col = 0
            for idx, header in enumerate(headers, start=1):
                if header not in (None, ""):
                    max_col = max(max_col, idx)
            for row in data_rows:
                for idx, cell in enumerate(row, start=1):
                    if cell not in (None, ""):
                        max_col = max(max_col, idx)
            if max_col == 0:
                max_col = max(len(headers), 1)
            columns = []
            for col_idx in range(1, max_col + 1):
                letter = get_column_letter(col_idx) if get_column_letter else str(col_idx)
                columns.append(letter)
            self.preview_tree["columns"] = columns
            for col_idx, letter in enumerate(columns, start=1):
                header_text = ""
                if col_idx - 1 < len(headers):
                    header_val = headers[col_idx - 1]
                    header_text = str(header_val).strip() if header_val is not None else ""
                heading = f"{letter} - {header_text}" if header_text else letter
                self.preview_tree.heading(letter, text=heading)
                self.preview_tree.column(letter, width=140, anchor="w", stretch=False)
            for row in data_rows:
                values = []
                for col_idx in range(1, max_col + 1):
                    value = row[col_idx - 1] if col_idx - 1 < len(row) else ""
                    values.append("" if value is None else str(value))
                self.preview_tree.insert("", tk.END, values=values)
        finally:
            workbook.close()

    # ------------------------------------------------------------------ Lock overlay
    def _show_lock_overlay(self) -> None:
        if not self._locked or self._lock_overlay is not None:
            return
        overlay = tk.Frame(self, bg="#111219")
        overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        self._lock_overlay = overlay
        card = ttk.Frame(overlay, padding=24)
        card.place(relx=0.5, rely=0.5, anchor="center")
        ttk.Label(
            card,
            text="Production Log tab is under development.",
            style="SidebarHeading.TLabel",
            wraplength=420,
            justify="center",
        ).pack(anchor="center")
        ttk.Label(
            card,
            text=(
                "For tracking production counts, updating the client's log spreadsheet, "
                "and keeping consistent formatting of the spreadsheet."
            ),
            wraplength=420,
            justify="center",
        ).pack(anchor="center", pady=(12, 20))
        ttk.Label(card, text="Enter PIN to unlock:", justify="center").pack(anchor="center")
        self._pin_var.set("")
        validate = (self.register(self._validate_pin), "%P")
        entry = ttk.Entry(
            card,
            show="*",
            textvariable=self._pin_var,
            justify="center",
            width=12,
            validate="key",
            validatecommand=validate,
        )
        entry.pack(anchor="center", pady=(6, 0))
        entry.bind("<Return>", self._attempt_unlock)
        entry.focus_set()
        self._pin_entry = entry
        ttk.Button(card, text="Unlock", command=self._attempt_unlock).pack(anchor="center", pady=(10, 0))
        ttk.Label(card, textvariable=self._lock_error_var, foreground="#F36C6C").pack(anchor="center", pady=(8, 0))

    def _validate_pin(self, proposed: str) -> bool:
        if not proposed:
            return True
        if not proposed.isdigit():
            return False
        return len(proposed) <= len(self._PIN_CODE)

    def _attempt_unlock(self, event: Optional[tk.Event] = None) -> Optional[str]:
        value = self._pin_var.get()
        if value == self._PIN_CODE:
            self._unlock()
            return "break"
        self._lock_error_var.set("Incorrect PIN. Try again.")
        self._pin_var.set("")
        if self._pin_entry is not None:
            self._pin_entry.focus_set()
        return "break"

    def _unlock(self) -> None:
        self._locked = False
        if self._lock_overlay is not None:
            self._lock_overlay.destroy()
            self._lock_overlay = None
        self._lock_error_var.set("")

    def apply_theme(self, theme: ThemePalette) -> None:
        self.theme = theme
        self._configure_styles()
        if hasattr(self, "_accent_strip") and self._accent_strip is not None:
            self._accent_strip.configure(bg=self.theme.accent)
        for widget in self.winfo_children():
            try:
                widget.configure(style="ProdLog.Root.TFrame")
            except tk.TclError:
                pass

    def _configure_styles(self) -> None:
        style = ttk.Style(self)
        style.configure("ProdLog.Root.TFrame", background=self.theme.surface_bg)
        style.configure("ProdLog.Hero.TFrame", background=self.theme.card_alt_bg)
        style.configure("ProdLog.Card.TFrame", background=self.theme.card_bg)
        style.configure(
            "ProdLog.Title.TLabel",
            background=self.theme.card_alt_bg,
            foreground=self.theme.text_primary,
            font=("Segoe UI", 14, "bold"),
        )
        style.configure(
            "ProdLog.Section.TLabel",
            background=self.theme.card_bg,
            foreground=self.theme.accent,
            font=("Segoe UI", 11, "bold"),
        )
        style.configure(
            "ProdLog.Body.TLabel",
            background=self.theme.card_alt_bg,
            foreground=self.theme.text_secondary,
            font=("Segoe UI", 10),
        )
        style.configure(
            "ProdLog.BodyMuted.TLabel",
            background=self.theme.card_bg,
            foreground=self.theme.text_muted,
            font=("Segoe UI", 9),
        )
        style.configure(
            "ProdLog.Card.TLabel",
            background=self.theme.card_bg,
            foreground=self.theme.text_primary,
            font=("Segoe UI", 10),
        )
        style.configure(
            "ProdLog.Badge.TLabel",
            background=self.theme.surface_alt_bg,
            foreground=self.theme.text_secondary,
            padding=(12, 4),
            font=("Segoe UI", 9, "bold"),
        )
        style.configure(
            "ProdLog.Treeview",
            background=self.theme.list_bg,
            fieldbackground=self.theme.list_bg,
            foreground=self.theme.text_primary,
            borderwidth=0,
            font=("Segoe UI", 10),
        )
        style.configure(
            "ProdLog.Treeview.Heading",
            background=self.theme.list_alt_bg,
            foreground=self.theme.text_secondary,
            font=("Segoe UI", 10, "bold"),
        )
        style.map(
            "ProdLog.Treeview",
            background=[("selected", self.theme.list_selected_bg)],
            foreground=[("selected", self.theme.list_selected_fg)],
        )

    def is_locked(self) -> bool:
        return self._locked

    def focus_lock_entry(self) -> None:
        if self._pin_entry is not None:
            self._pin_entry.focus_set()

    def notify_locked(self) -> None:
        messagebox.showinfo("Production Log", "Enter the PIN to unlock this tab.", parent=self)
        self.focus_lock_entry()
