from __future__ import annotations

import calendar as cal
import hashlib
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from textwrap import shorten
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import colorchooser, filedialog, messagebox, simpledialog, ttk

from .database import Database
from .models import IssueClient, IssueItem, IssueNote, IssuePublication
from . import utils
from .theme import ThemePalette

WEEKDAY_NAMES = ["Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat"]


@dataclass
class DayCell:
    frame: tk.Frame
    day_label: tk.Label
    events_container: tk.Frame
    date: Optional[date] = None


@dataclass
class IssueOccurrence:
    item: IssueItem
    run_type: str
    run_date: date


class IssueCalendarTab(ttk.Frame):
    def __init__(self, master: tk.Misc, db: Database, theme: ThemePalette) -> None:
        super().__init__(master)
        self.db = db
        self.theme = theme
        self.configure(padding=(16, 16))

        self.current_month = datetime.now().date().replace(day=1)
        self.selected_day = datetime.now().date()

        self.clients: List[IssueClient] = []
        self.current_client_id: Optional[int] = None
        self.items: List[IssueItem] = []
        self.publications: List[IssuePublication] = []
        self.visible_publications: set[str] = set()
        self.publication_vars: Dict[str, tk.BooleanVar] = {}
        self._publication_checkbuttons: List[ttk.Checkbutton] = []
        self._publication_color_canvases: List[tk.Canvas] = []
        self.occurrences_by_day: Dict[date, List[IssueOccurrence]] = {}
        self.day_cells: List[DayCell] = []
        self.selected_cell: Optional[DayCell] = None
        self._day_occurrence_index: Dict[str, IssueOccurrence] = {}
        self._detail_overlay: Optional[tk.Frame] = None
        self._detail_panel: Optional[tk.Frame] = None
        self._last_import_backup: Optional[List[dict]] = None
        self._last_import_client_id: Optional[int] = None
        self.undo_import_button: Optional[ttk.Button] = None
        self.publications_frame: Optional[ttk.Frame] = None

        self._assign_palette_colors()
        self._build_ui()
        self.refresh()

    def _assign_palette_colors(self) -> None:
        palette = self.theme
        self.bg_color = palette.surface_bg
        self.cell_bg = palette.calendar_cell_bg
        self.cell_selected_bg = palette.calendar_cell_selected_bg
        self.outside_month_color = palette.calendar_outside_text
        self.text_color = palette.text_primary
        self.secondary_text_color = palette.text_secondary
        self.list_bg = palette.list_bg
        self.list_fg = palette.text_primary
        self.list_selected_bg = palette.list_selected_bg
        self.list_selected_fg = palette.list_selected_fg

    def apply_theme(self, theme: ThemePalette) -> None:
        self.theme = theme
        self._assign_palette_colors()
        self._build_ui()
        self.refresh()

    # ------------------------------------------------------------------ UI
    def _build_ui(self) -> None:
        for child in self.winfo_children():
            child.destroy()

        selector = ttk.Frame(self)
        selector.pack(fill=tk.X, pady=(0, 12))

        ttk.Label(selector, text="Issue Calendar", style="SidebarHeading.TLabel").pack(side=tk.LEFT)
        self.client_combo = ttk.Combobox(selector, state="readonly", width=30)
        self.client_combo.pack(side=tk.LEFT, padx=(12, 0))
        self.client_combo.bind("<<ComboboxSelected>>", self._on_client_selected)

        ttk.Button(selector, text="New...", command=self.add_client).pack(side=tk.LEFT, padx=(12, 0))
        ttk.Button(selector, text="Rename...", command=self.rename_client).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(selector, text="Delete", command=self.delete_client).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(selector, text="Import...", command=self.import_issue_calendar).pack(side=tk.LEFT, padx=(12, 0))
        self.undo_import_button = ttk.Button(
            selector,
            text="Undo Import",
            command=self.undo_last_import,
            state=tk.DISABLED,
        )
        self.undo_import_button.pack(side=tk.LEFT, padx=(6, 0))

        paned = ttk.Panedwindow(self, orient=tk.HORIZONTAL)
        paned.pack(fill=tk.BOTH, expand=True)

        container = ttk.Frame(paned)
        container.columnconfigure(0, weight=1)
        container.rowconfigure(0, weight=1)

        sidebar_outer = ttk.Frame(paned)
        sidebar_outer.columnconfigure(0, weight=1)
        sidebar_outer.rowconfigure(0, weight=1)

        paned.add(container, weight=3)
        paned.add(sidebar_outer, weight=2)

        # Left: calendar grid --------------------------------------------------
        left = ttk.Frame(container)
        left.grid(row=0, column=0, sticky="nsew")
        left.columnconfigure(0, weight=1)
        left.rowconfigure(1, weight=1)

        toolbar = ttk.Frame(left)
        toolbar.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        toolbar.columnconfigure(2, weight=1)

        ttk.Button(toolbar, text="<", width=3, command=self.go_to_previous_month).grid(row=0, column=0, padx=(0, 6))
        ttk.Button(toolbar, text=">", width=3, command=self.go_to_next_month).grid(row=0, column=1, padx=(0, 6))
        self.month_label = ttk.Label(toolbar, text="", style="CalendarHeading.TLabel")
        self.month_label.grid(row=0, column=2)
        ttk.Button(toolbar, text="Today", command=self.go_to_today).grid(row=0, column=3, padx=6)

        grid_frame = ttk.Frame(left)
        grid_frame.grid(row=1, column=0, sticky="nsew")
        for c in range(7):
            grid_frame.columnconfigure(c, weight=1, uniform="day")
        for r in range(6):
            grid_frame.rowconfigure(r + 1, weight=1, uniform="dayrow")

        for col, name in enumerate(WEEKDAY_NAMES):
            header = tk.Label(
                grid_frame,
                text=name,
                bg=self.bg_color,
                fg=self.secondary_text_color,
                padx=4,
                pady=4,
                font=("Segoe UI", 10, "bold"),
            )
            header.grid(row=0, column=col, sticky="nsew", padx=1, pady=1)

        self.day_cells = []
        for row in range(6):
            for col in range(7):
                frame = tk.Frame(grid_frame, bg=self.cell_bg, bd=0, highlightthickness=0)
                frame.grid(row=row + 1, column=col, sticky="nsew", padx=1, pady=1)
                frame.bind("<Button-1>", lambda e, idx=len(self.day_cells): self._on_cell_click(idx))

                day_label = tk.Label(
                    frame,
                    text="",
                    anchor="nw",
                    bg=self.cell_bg,
                    fg=self.text_color,
                    font=("Segoe UI", 11, "bold"),
                    padx=6,
                    pady=4,
                )
                day_label.pack(fill=tk.X)
                day_label.bind("<Button-1>", lambda e, idx=len(self.day_cells): self._on_cell_click(idx))

                events_container = tk.Frame(frame, bg=self.cell_bg)
                events_container.pack(fill=tk.BOTH, expand=True, padx=4, pady=(0, 4))
                events_container.bind("<Button-1>", lambda e, idx=len(self.day_cells): self._on_cell_click(idx))

                self.day_cells.append(DayCell(frame=frame, day_label=day_label, events_container=events_container))

        sidebar = ttk.Frame(sidebar_outer, padding=(12, 0))
        sidebar.grid(row=0, column=0, sticky="nsew")
        sidebar_outer.columnconfigure(0, weight=1)
        sidebar_outer.rowconfigure(0, weight=1)

        sidebar.columnconfigure(0, weight=1)
        sidebar.rowconfigure(0, weight=0)
        sidebar.rowconfigure(1, weight=1, minsize=16)
        sidebar.rowconfigure(2, weight=0)

        top_container = ttk.Frame(sidebar)
        top_container.grid(row=0, column=0, sticky="ew")
        top_container.columnconfigure(0, weight=1)

        publications_label = ttk.Label(top_container, text="Publications", style="SidebarHeading.TLabel")
        publications_label.grid(row=0, column=0, sticky="w")

        self.publications_frame = ttk.Frame(top_container)
        self.publications_frame.grid(row=1, column=0, sticky="ew", pady=(6, 12))
        self.publications_frame.columnconfigure(1, weight=1)

        selected_container = ttk.Frame(sidebar, padding=(0, 0))
        selected_container.grid(row=2, column=0, sticky="sew", pady=(24, 0))
        selected_container.columnconfigure(0, weight=1)
        selected_container.rowconfigure(2, weight=1, minsize=260)

        ttk.Label(selected_container, text="Selected Day", style="SidebarHeading.TLabel").grid(
            row=0, column=0, sticky="w"
        )
        self.day_value_label = ttk.Label(selected_container, text="", style="SelectedDay.TLabel")
        self.day_value_label.grid(row=1, column=0, sticky="w", pady=(0, 8))

        self.day_events_tree = ttk.Treeview(
            selected_container,
            columns=("pub", "run", "issue"),
            show="headings",
            height=10,
        )
        self.day_events_tree.heading("pub", text="Pub")
        self.day_events_tree.heading("run", text="Run")
        self.day_events_tree.heading("issue", text="Issue")
        self.day_events_tree.column("pub", width=80, anchor="w")
        self.day_events_tree.column("run", width=70, anchor="center")
        self.day_events_tree.column("issue", width=200, anchor="w")
        self.day_events_tree.grid(row=2, column=0, sticky="nsew")
        self.day_events_tree.bind("<Double-1>", self._open_selected_occurrence)

        scroll = ttk.Scrollbar(selected_container, orient=tk.VERTICAL, command=self.day_events_tree.yview)
        scroll.grid(row=2, column=1, sticky="ns")
        self.day_events_tree.configure(yscrollcommand=scroll.set)

    # ------------------------------------------------------------------ Data flow
    def refresh(self) -> None:
        self._load_clients()
        self._load_items()
        self._populate_calendar()
        self._update_day_details()

    def _load_clients(self) -> None:
        self.clients = self.db.get_issue_clients()
        names = [client.name for client in self.clients]
        self.client_combo["values"] = names
        if self.current_client_id and any(c.id == self.current_client_id for c in self.clients):
            current = next(c for c in self.clients if c.id == self.current_client_id)
            self.client_combo.set(current.name)
            return
        if names:
            self.client_combo.current(0)
            self.current_client_id = self.clients[0].id
        else:
            self.client_combo.set("")
            self.current_client_id = None
        self._sync_undo_state()

    def _load_items(self) -> None:
        if self.current_client_id is None:
            self.items = []
            self.publications = []
            self.visible_publications = set()
            self.occurrences_by_day = {}
            self._rebuild_publication_filters()
            return
        self.items = self.db.get_issue_items(self.current_client_id)
        self._load_publications()
        self.occurrences_by_day = self._build_occurrences(self.items)

    def _build_occurrences(self, items: List[IssueItem]) -> Dict[date, List[IssueOccurrence]]:
        occurrences: Dict[date, List[IssueOccurrence]] = {}
        for item in items:
            if self.publications and item.publication_code not in self.visible_publications:
                continue
            if item.trial_date:
                occurrences.setdefault(item.trial_date, []).append(
                    IssueOccurrence(item=item, run_type="Trial", run_date=item.trial_date)
                )
            if item.update_date:
                occurrences.setdefault(item.update_date, []).append(
                    IssueOccurrence(item=item, run_type="Update", run_date=item.update_date)
                )
        for day in occurrences:
            occurrences[day].sort(key=lambda occ: (occ.item.publication_code.lower(), occ.item.issue_name.lower()))
        return occurrences

    def _load_publications(self) -> None:
        if self.current_client_id is None:
            self.publications = []
            self.visible_publications = set()
            self._rebuild_publication_filters()
            return
        codes = sorted({item.publication_code for item in self.items if item.publication_code})
        existing = {pub.publication_code: pub for pub in self.db.get_issue_publications(self.current_client_id)}
        for code in codes:
            if code not in existing:
                color = self._color_for_publication(code)
                try:
                    self.db.upsert_issue_publication(
                        client_id=self.current_client_id,
                        publication_code=code,
                        color=color,
                        is_visible=True,
                    )
                except ValueError:
                    continue
        publications = self.db.get_issue_publications(self.current_client_id)
        self.publications = [pub for pub in publications if pub.publication_code in codes]
        self.visible_publications = {pub.publication_code for pub in self.publications if pub.is_visible}
        self._rebuild_publication_filters()

    def _populate_calendar(self) -> None:
        month_start = self.current_month
        cal_obj = cal.Calendar(firstweekday=6)
        weeks = cal_obj.monthdatescalendar(month_start.year, month_start.month)

        if self.month_label:
            self.month_label.configure(text=month_start.strftime("%B %Y"))

        while len(weeks) < 6:
            weeks.append([day + timedelta(days=7) for day in weeks[-1]])

        for cell in self.day_cells:
            for widget in cell.events_container.winfo_children():
                widget.destroy()

        for idx, cell in enumerate(self.day_cells):
            row = idx // 7
            col = idx % 7
            day = weeks[row][col]
            cell.date = day
            is_current = day.month == month_start.month
            cell.day_label.configure(text=str(day.day))
            fg_color = self.text_color if is_current else self.outside_month_color
            cell.day_label.configure(fg=fg_color)
            cell.frame.configure(bg=self.cell_bg)
            cell.day_label.configure(bg=self.cell_bg)
            cell.events_container.configure(bg=self.cell_bg)

            events = self.occurrences_by_day.get(day, [])
            for occ in events[:4]:
                color = self._color_for_publication(occ.item.publication_code)
                label_text = shorten(occ.item.issue_name, width=20, placeholder="...")
                event_label = tk.Label(
                    cell.events_container,
                    text=label_text,
                    bg=color,
                    fg=self._text_color_for_bg(color),
                    padx=4,
                    pady=1,
                    anchor="w",
                    font=("Segoe UI", 8, "bold"),
                )
                event_label.pack(fill=tk.X, pady=1)
                event_label.bind("<Double-1>", lambda _e, occ=occ: self._open_issue_detail(occ))
            if len(events) > 4:
                more_label = tk.Label(
                    cell.events_container,
                    text=f"+{len(events) - 4} more",
                    bg=self.cell_bg,
                    fg=self.secondary_text_color,
                    anchor="w",
                    font=("Segoe UI", 8),
                )
                more_label.pack(fill=tk.X, pady=(2, 0))

        self._select_day_cell(self.selected_day)

    def _select_day_cell(self, day: date) -> None:
        for cell in self.day_cells:
            if cell.date == day:
                self.selected_cell = cell
                cell.frame.configure(bg=self.cell_selected_bg)
                cell.day_label.configure(bg=self.cell_selected_bg)
                cell.events_container.configure(bg=self.cell_selected_bg)
            else:
                cell.frame.configure(bg=self.cell_bg)
                cell.day_label.configure(bg=self.cell_bg)
                cell.events_container.configure(bg=self.cell_bg)

    def _update_day_details(self) -> None:
        if not self.day_events_tree or not self.day_value_label:
            return
        self.day_value_label.configure(text=self.selected_day.strftime("%B %d, %Y"))
        for item in self.day_events_tree.get_children():
            self.day_events_tree.delete(item)
        self._day_occurrence_index.clear()

        occurrences = self.occurrences_by_day.get(self.selected_day, [])
        for occ in occurrences:
            item_id = self.day_events_tree.insert(
                "",
                tk.END,
                values=(occ.item.publication_code, occ.run_type, occ.item.issue_name),
            )
            self._day_occurrence_index[item_id] = occ

    def _rebuild_publication_filters(self) -> None:
        if self.publications_frame is None:
            return
        for child in self.publications_frame.winfo_children():
            child.destroy()
        self.publication_vars.clear()
        self._publication_checkbuttons = []
        self._publication_color_canvases = []

        for pub in self.publications:
            row = ttk.Frame(self.publications_frame)
            row.pack(fill=tk.X, pady=(2, 6))
            color_value = pub.color or self._color_for_publication(pub.publication_code)

            patch = tk.Canvas(row, width=18, height=18, highlightthickness=0, bg=self.bg_color)
            patch.create_rectangle(0, 0, 18, 18, fill=color_value, outline="")
            patch.pack(side=tk.LEFT, padx=(0, 6))

            var = tk.BooleanVar(value=pub.is_visible)
            check = ttk.Checkbutton(
                row,
                text=pub.publication_code,
                variable=var,
                command=lambda code=pub.publication_code, v=var: self._toggle_publication(code, v.get()),
            )
            check.pack(side=tk.LEFT)

            edit_btn = ttk.Button(
                row,
                text="Edit",
                command=lambda code=pub.publication_code: self._open_publication_editor(code),
                width=8,
            )
            edit_btn.pack(side=tk.RIGHT)

            self.publication_vars[pub.publication_code] = var
            self._publication_checkbuttons.append(check)
            self._publication_color_canvases.append(patch)

        if not self.publications:
            ttk.Label(self.publications_frame, text="No publications yet.").pack(anchor="w")

    def _toggle_publication(self, code: str, visible: bool) -> None:
        if self.current_client_id is None:
            return
        try:
            self.db.update_issue_publication(
                client_id=self.current_client_id,
                publication_code=code,
                is_visible=visible,
            )
        except ValueError:
            return
        if visible:
            self.visible_publications.add(code)
        else:
            self.visible_publications.discard(code)
        self.occurrences_by_day = self._build_occurrences(self.items)
        self._populate_calendar()
        self._update_day_details()

    def _open_publication_editor(self, code: str) -> None:
        if self.current_client_id is None:
            return
        publication = next((p for p in self.publications if p.publication_code == code), None)
        if publication is None:
            return
        panel = tk.Toplevel(self)
        panel.title("Edit Publication")
        panel.transient(self.winfo_toplevel())
        panel.resizable(False, False)
        container = ttk.Frame(panel, padding=16)
        container.pack(fill=tk.BOTH, expand=True)
        container.columnconfigure(1, weight=1)

        ttk.Label(container, text="Publication Code:", style="SidebarHeading.TLabel").grid(row=0, column=0, sticky="w")
        ttk.Label(container, text=publication.publication_code).grid(row=0, column=1, sticky="w", padx=(8, 0))

        ttk.Label(container, text="Color:").grid(row=1, column=0, sticky="w", pady=(12, 0))
        color_preview = tk.Canvas(container, width=28, height=18, highlightthickness=0, bg=self.bg_color)
        color_preview.grid(row=1, column=1, sticky="w", pady=(12, 0))
        color_value = publication.color or self._color_for_publication(publication.publication_code)
        rect = color_preview.create_rectangle(0, 0, 28, 18, fill=color_value, outline="")

        def pick_color() -> None:
            current = publication.color or self._color_for_publication(publication.publication_code)
            chosen = colorchooser.askcolor(initialcolor=current, title=f"Pick color for {code}")
            if not chosen or not chosen[1]:
                return
            new_color = chosen[1]
            try:
                self.db.update_issue_publication(
                    client_id=self.current_client_id,
                    publication_code=publication.publication_code,
                    color=new_color,
                )
            except ValueError:
                return
            publication.color = new_color
            color_preview.itemconfigure(rect, fill=new_color)
            self._rebuild_publication_filters()
            self._populate_calendar()

        ttk.Button(container, text="Edit Color", command=pick_color).grid(row=1, column=2, padx=(8, 0), pady=(12, 0))

        ttk.Button(container, text="Close", command=panel.destroy).grid(row=2, column=2, sticky="e", pady=(16, 0))

    def _edit_publication_color(self, code: str) -> None:
        # Deprecated: use _open_publication_editor.
        self._open_publication_editor(code)

    # ------------------------------------------------------------------ Interactions
    def _on_cell_click(self, idx: int) -> None:
        if idx < 0 or idx >= len(self.day_cells):
            return
        cell = self.day_cells[idx]
        if cell.date is None:
            return
        self.selected_day = cell.date
        self._select_day_cell(self.selected_day)
        self._update_day_details()

    def go_to_previous_month(self) -> None:
        prev_month = utils.add_months(datetime.combine(self.current_month, datetime.min.time()), -1).date()
        self.current_month = prev_month.replace(day=1)
        if self.selected_day.month != self.current_month.month or self.selected_day.year != self.current_month.year:
            self.selected_day = self.current_month
        self._populate_calendar()
        self._update_day_details()

    def go_to_next_month(self) -> None:
        next_month = utils.add_months(datetime.combine(self.current_month, datetime.min.time()), 1).date()
        self.current_month = next_month.replace(day=1)
        if self.selected_day.month != self.current_month.month or self.selected_day.year != self.current_month.year:
            self.selected_day = self.current_month
        self._populate_calendar()
        self._update_day_details()

    def go_to_today(self) -> None:
        today = datetime.now().date()
        self.current_month = today.replace(day=1)
        self.selected_day = today
        self._populate_calendar()
        self._update_day_details()

    def _on_client_selected(self, _event: object) -> None:
        selected = self.client_combo.get()
        match = next((c for c in self.clients if c.name == selected), None)
        self.current_client_id = match.id if match else None
        self._load_items()
        self._populate_calendar()
        self._update_day_details()
        self._sync_undo_state()

    # ------------------------------------------------------------------ Client management
    def add_client(self) -> None:
        name = simpledialog.askstring("New Client", "Client name:", parent=self)
        if not name:
            return
        try:
            client_id = self.db.create_issue_client(name)
        except ValueError as exc:
            messagebox.showerror("New Client", str(exc), parent=self)
            return
        self.current_client_id = client_id
        self.refresh()

    def rename_client(self) -> None:
        if self.current_client_id is None:
            messagebox.showinfo("Rename Client", "Select a client first.", parent=self)
            return
        current = next((c for c in self.clients if c.id == self.current_client_id), None)
        if current is None:
            return
        name = simpledialog.askstring("Rename Client", "Client name:", initialvalue=current.name, parent=self)
        if not name:
            return
        try:
            self.db.update_issue_client(current.id, name)
        except ValueError as exc:
            messagebox.showerror("Rename Client", str(exc), parent=self)
            return
        self.refresh()

    def delete_client(self) -> None:
        if self.current_client_id is None:
            messagebox.showinfo("Delete Client", "Select a client first.", parent=self)
            return
        client = next((c for c in self.clients if c.id == self.current_client_id), None)
        if client is None:
            return
        if not messagebox.askyesno(
            "Delete Client",
            f"Delete client '{client.name}' and all of its issue items?",
            parent=self,
        ):
            return
        self.db.delete_issue_client(client.id)
        self.current_client_id = None
        if self._last_import_client_id == client.id:
            self._last_import_client_id = None
            self._last_import_backup = None
        self.refresh()

    # ------------------------------------------------------------------ Import
    def import_issue_calendar(self) -> None:
        if self.current_client_id is None:
            messagebox.showinfo("Issue Calendar", "Select a client before importing.", parent=self)
            return
        path = filedialog.askopenfilename(
            parent=self,
            title="Import Issue Calendar",
            filetypes=[("Excel files", "*.xlsx;*.xlsm;*.xltx;*.xltm"), ("All files", "*.*")],
        )
        if not path:
            return
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            messagebox.showerror(
                "Missing Dependency",
                "openpyxl is required to import Excel files. Please install it and try again.",
                parent=self,
            )
            return
        try:
            workbook = load_workbook(path, data_only=True)
            sheet = workbook.active
        except Exception as exc:
            messagebox.showerror("Import Failed", f"Could not read Excel file: {exc}", parent=self)
            return

        import_date = datetime.now().date()
        backup = self._snapshot_current_client()
        added = 0
        updated = 0
        notes_added = 0
        skipped = 0

        for row in sheet.iter_rows(min_row=5, values_only=True):
            if not row or len(row) < 5:
                continue
            publication_code = self._clean_text(row[0])
            issue_name = self._clean_text(row[1])
            issue_number = self._clean_text(row[2])
            trial_date = self._parse_excel_date(row[3])
            update_date = self._parse_excel_date(row[4])
            note_text = self._clean_text(row[5]) if len(row) > 5 else ""

            if not publication_code or not issue_name:
                skipped += 1
                continue

            if trial_date and trial_date < import_date:
                trial_date = None
            if update_date and update_date < import_date:
                update_date = None

            if trial_date is None and update_date is None:
                skipped += 1
                continue

            existing = self.db.find_issue_item(self.current_client_id, publication_code, issue_name)
            item_id = self.db.upsert_issue_item(
                client_id=self.current_client_id,
                publication_code=publication_code,
                issue_name=issue_name,
                issue_number=issue_number or None,
                trial_date=trial_date,
                update_date=update_date,
            )
            if existing:
                updated += 1
            else:
                added += 1

            if note_text:
                existing_notes = self.db.get_issue_notes(item_id)
                if all(note.content != note_text for note in existing_notes):
                    try:
                        self.db.add_issue_note(item_id, note_text)
                        notes_added += 1
                    except ValueError:
                        pass

        self._load_items()
        self._populate_calendar()
        self._update_day_details()
        self._last_import_backup = backup
        self._last_import_client_id = self.current_client_id
        self._sync_undo_state()
        messagebox.showinfo(
            "Import Complete",
            f"Added {added}, updated {updated}, notes added {notes_added}, skipped {skipped}.",
            parent=self,
        )

    @staticmethod
    def _clean_text(value: object) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        return text

    @staticmethod
    def _parse_excel_date(value: object) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                continue
        return None

    def _snapshot_current_client(self) -> Optional[List[dict]]:
        if self.current_client_id is None:
            return None
        snapshot: List[dict] = []
        for item in self.db.get_issue_items(self.current_client_id):
            notes = self.db.get_issue_notes(item.id)
            snapshot.append(
                {
                    "publication_code": item.publication_code,
                    "issue_name": item.issue_name,
                    "issue_number": item.issue_number,
                    "trial_date": item.trial_date.isoformat() if item.trial_date else None,
                    "update_date": item.update_date.isoformat() if item.update_date else None,
                    "created_at": utils.to_iso(item.created_at),
                    "updated_at": utils.to_iso(item.updated_at) if item.updated_at else None,
                    "notes": [
                        {
                            "content": note.content,
                            "created_at": utils.to_iso(note.created_at),
                            "updated_at": utils.to_iso(note.updated_at) if note.updated_at else None,
                        }
                        for note in notes
                    ],
                }
            )
        return snapshot

    def undo_last_import(self) -> None:
        if (
            self.current_client_id is None
            or self._last_import_client_id != self.current_client_id
            or self._last_import_backup is None
        ):
            messagebox.showinfo("Undo Import", "There is no import to undo for this client.", parent=self)
            self._sync_undo_state()
            return
        if not messagebox.askyesno(
            "Undo Import",
            "Restore the issue calendar to its state before the last import?",
            parent=self,
        ):
            return
        self.db.replace_issue_client_data(self.current_client_id, self._last_import_backup)
        self._last_import_backup = None
        self._last_import_client_id = None
        self._load_items()
        self._populate_calendar()
        self._update_day_details()
        self._sync_undo_state()

    def _sync_undo_state(self) -> None:
        if self.undo_import_button is None:
            return
        enabled = (
            self.current_client_id is not None
            and self._last_import_backup is not None
            and self._last_import_client_id == self.current_client_id
        )
        self.undo_import_button.configure(state=tk.NORMAL if enabled else tk.DISABLED)

    # ------------------------------------------------------------------ Details & notes
    def _open_selected_occurrence(self, _event: object) -> None:
        selection = self.day_events_tree.selection()
        if not selection:
            return
        occ = self._day_occurrence_index.get(selection[0])
        if not occ:
            return
        self._open_issue_detail(occ)

    def _open_issue_detail(self, occurrence: IssueOccurrence) -> None:
        self._close_detail_panel()
        overlay = tk.Frame(self, bg=self.bg_color)
        overlay.place(relx=0, rely=0, relwidth=1, relheight=1)
        overlay.lift()
        panel = tk.Frame(overlay, bg=self.theme.card_bg, bd=1, relief="ridge")
        panel.place(relx=0.5, rely=0.5, anchor="center")
        panel.columnconfigure(0, weight=1)

        def close() -> None:
            self._close_detail_panel()

        container = ttk.Frame(panel, padding=16)
        container.grid(row=0, column=0, sticky="nsew")
        container.columnconfigure(0, weight=1)

        ttk.Label(container, text=occurrence.item.issue_name, style="SidebarHeading.TLabel").pack(anchor="w")
        ttk.Label(
            container,
            text=f"Publication: {occurrence.item.publication_code}",
        ).pack(anchor="w", pady=(4, 0))
        issue_number = occurrence.item.issue_number or "—"
        ttk.Label(container, text=f"Issue number: {issue_number}").pack(anchor="w", pady=(2, 0))
        trial_text = occurrence.item.trial_date.strftime("%m/%d/%Y") if occurrence.item.trial_date else "—"
        update_text = occurrence.item.update_date.strftime("%m/%d/%Y") if occurrence.item.update_date else "—"
        ttk.Label(container, text=f"Trial run: {trial_text}").pack(anchor="w", pady=(4, 0))
        ttk.Label(container, text=f"Update run: {update_text}").pack(anchor="w", pady=(2, 10))

        notes_frame = ttk.Frame(container)
        notes_frame.pack(fill=tk.BOTH, expand=True)
        notes_frame.columnconfigure(0, weight=1)

        ttk.Label(notes_frame, text="Notes", style="SidebarHeading.TLabel").grid(row=0, column=0, sticky="w")
        notes_tree = ttk.Treeview(
            notes_frame,
            columns=("when", "note"),
            show="headings",
            height=8,
        )
        notes_tree.heading("when", text="When")
        notes_tree.heading("note", text="Note")
        notes_tree.column("when", width=140, anchor="w")
        notes_tree.column("note", width=360, anchor="w")
        notes_tree.grid(row=1, column=0, sticky="nsew", pady=(6, 0))
        notes_frame.rowconfigure(1, weight=1)
        scroll = ttk.Scrollbar(notes_frame, orient=tk.VERTICAL, command=notes_tree.yview)
        scroll.grid(row=1, column=1, sticky="ns")
        notes_tree.configure(yscrollcommand=scroll.set)

        note_rows = self.db.get_issue_notes(occurrence.item.id)
        note_index: Dict[str, IssueNote] = {}
        for note in note_rows:
            when = note.updated_at or note.created_at
            when_str = when.strftime("%Y-%m-%d %H:%M")
            entry_id = notes_tree.insert("", tk.END, values=(when_str, note.content))
            note_index[entry_id] = note

        def add_note() -> None:
            text = simpledialog.askstring("New Note", "Note:", parent=panel)
            if not text:
                return
            try:
                note_id = self.db.add_issue_note(occurrence.item.id, text)
            except ValueError as exc:
                messagebox.showerror("Note", str(exc), parent=panel)
                return
            now = datetime.now().strftime("%Y-%m-%d %H:%M")
            entry_id = notes_tree.insert("", tk.END, values=(now, text))
            note_index[entry_id] = IssueNote(
                id=note_id,
                item_id=occurrence.item.id,
                content=text.strip(),
                created_at=datetime.now(),
                updated_at=datetime.now(),
            )

        def edit_note() -> None:
            selection = notes_tree.selection()
            if not selection:
                return
            note = note_index.get(selection[0])
            if note is None:
                return
            text = simpledialog.askstring("Edit Note", "Note:", initialvalue=note.content, parent=panel)
            if not text:
                return
            try:
                self.db.update_issue_note(note.id, text)
            except ValueError as exc:
                messagebox.showerror("Note", str(exc), parent=panel)
                return
            note.content = text.strip()
            note.updated_at = datetime.now()
            notes_tree.item(selection[0], values=(note.updated_at.strftime("%Y-%m-%d %H:%M"), note.content))

        def delete_note() -> None:
            selection = notes_tree.selection()
            if not selection:
                return
            note = note_index.get(selection[0])
            if note is None:
                return
            if not messagebox.askyesno("Delete Note", "Delete this note?", parent=panel):
                return
            self.db.delete_issue_note(note.id)
            notes_tree.delete(selection[0])
            note_index.pop(selection[0], None)

        note_buttons = ttk.Frame(notes_frame)
        note_buttons.grid(row=2, column=0, sticky="w", pady=(6, 0))
        ttk.Button(note_buttons, text="Add Note", command=add_note).pack(side=tk.LEFT)
        ttk.Button(note_buttons, text="Edit Note", command=edit_note).pack(side=tk.LEFT, padx=(6, 0))
        ttk.Button(note_buttons, text="Delete Note", command=delete_note).pack(side=tk.LEFT, padx=(6, 0))

        actions = ttk.Frame(container)
        actions.pack(fill=tk.X, pady=(12, 0))

        def delete_occurrence() -> None:
            if not messagebox.askyesno(
                "Delete Occurrence",
                f"Delete the {occurrence.run_type.lower()} run for this issue?",
                parent=panel,
            ):
                return
            if not messagebox.askyesno(
                "Confirm Delete",
                "This cannot be undone. Delete this occurrence?",
                parent=panel,
            ):
                return
            trial_date = occurrence.item.trial_date
            update_date = occurrence.item.update_date
            if occurrence.run_type == "Trial":
                trial_date = None
            else:
                update_date = None
            self.db.update_issue_item_dates(
                occurrence.item.id,
                trial_date=trial_date,
                update_date=update_date,
            )
            self._load_items()
            self._populate_calendar()
            self._update_day_details()
            close()

        ttk.Button(actions, text="Delete Occurrence", command=delete_occurrence, style="Danger.TButton").pack(
            side=tk.LEFT
        )
        ttk.Button(actions, text="Close", command=close).pack(side=tk.RIGHT)

        overlay.bind("<Button-1>", lambda _e: "break")
        overlay.bind("<Escape>", lambda _e: close())
        panel.focus_set()
        self._detail_overlay = overlay
        self._detail_panel = panel

    def _close_detail_panel(self) -> None:
        if self._detail_panel is not None:
            self._detail_panel.destroy()
            self._detail_panel = None
        if self._detail_overlay is not None:
            self._detail_overlay.destroy()
            self._detail_overlay = None

    # ------------------------------------------------------------------ Colors
    @staticmethod
    def _text_color_for_bg(color: str) -> str:
        text = color.lstrip("#")
        if len(text) != 6:
            return "#ffffff"
        r = int(text[0:2], 16)
        g = int(text[2:4], 16)
        b = int(text[4:6], 16)
        luminance = (0.2126 * r + 0.7152 * g + 0.0722 * b) / 255
        return "#000000" if luminance > 0.6 else "#ffffff"

    def _color_for_publication(self, code: str) -> str:
        current = next((p for p in self.publications if p.publication_code == code), None)
        if current and current.color:
            return current.color
        palette = [
            "#4F75FF",
            "#3BAA7D",
            "#D97757",
            "#B57EDC",
            "#F2B94E",
            "#6B9AC4",
            "#F06C9B",
            "#49A1A6",
            "#6F7CFF",
            "#8BC34A",
            "#E57373",
            "#9575CD",
        ]
        digest = hashlib.md5(code.encode("utf-8")).hexdigest()
        idx = int(digest, 16) % len(palette)
        return palette[idx]


__all__ = ["IssueCalendarTab"]
